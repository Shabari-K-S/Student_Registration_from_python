"""

@author : Shabari K S
Created as Mini Project

"""


from tkinter import *
from tkinter import messagebox
import tkinter
import customtkinter
from openpyxl import *



customtkinter.set_appearance_mode("dark")
customtkinter.set_default_color_theme("dark-blue")


app = customtkinter.CTk()
app.geometry("520x420")
app.title("Login")
app.resizable(False, False)

wb = load_workbook('{ Path of the excel sheet that already exist }')

sheet = wb.active
 
 
def excel():
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 30
    sheet.column_dimensions['C'].width = 50
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 10
    sheet.column_dimensions['F'].width = 10

    sheet.cell(row=1, column=1).value = "First Name"
    sheet.cell(row=1, column=2).value = "Last Name"
    sheet.cell(row=1, column=3).value = "Email ID"
    sheet.cell(row=1, column=4).value = "Contact Number"
    sheet.cell(row=1, column=5).value = "Department"
    sheet.cell(row=1, column=6).value = "Year"
 
 
def focus1(event):
    b1.focus_set()
def focus2(event):
    c1.focus_set()
def focus3(event):
    d1.focus_set()
def focus4(event):
    e1.focus_set()
def focus5(event):
    f1.focus_set()


def clear():
    a1.delete(0, END)
    b1.delete(0, END)
    c1.delete(0, END)
    d1.delete(0, END)


def insert():
	if ( a1.get() == "" and b1.get() == "" and c1.get() == "" and d1.get() == "" and e1.get() == "-Select-" and f1.get() == "-Select-" ):
		print("empty input")
	else:
		current_row = sheet.max_row
		current_column = sheet.max_column
		sheet.cell(row=current_row + 1, column=1).value = a1.get()
		sheet.cell(row=current_row + 1, column=2).value = b1.get()
		sheet.cell(row=current_row + 1, column=3).value = c1.get()
		sheet.cell(row=current_row + 1, column=4).value = d1.get()
		sheet.cell(row=current_row + 1, column=5).value = e1.get()
		sheet.cell(row=current_row + 1, column=6).value = f1.get()
		wb.save('{ Path of the excel sheet that already exist }')
		a1.focus_set()
		clear()



frame_left = customtkinter.CTkFrame(master=app, width = 500, height = 400)
frame_left.place(x=10,y=10)


label_2 = customtkinter.CTkLabel(master=frame_left, text="Student Registration From",text_font = ('Microsoft Yahei UI Light',20,"bold"),  justify=tkinter.LEFT)
label_2.place(x=100,y=5)


a = customtkinter.CTkLabel(master = frame_left ,text = "First Name", justify = tkinter.LEFT).place(x = 10,y = 50)
b = customtkinter.CTkLabel(master = frame_left ,text = "Last Name :", justify = tkinter.LEFT).place(x = 10,y = 100)
c = customtkinter.CTkLabel(master = frame_left ,text = "Email Id :", justify = tkinter.LEFT).place(x = 10,y = 150)
d = customtkinter.CTkLabel(master = frame_left ,text = "Contact Number :", justify = tkinter.LEFT).place(x = 10,y = 200)
e = customtkinter.CTkLabel(master = frame_left ,text = "Department :", justify = tkinter.LEFT).place(x = 10,y = 250)
f = customtkinter.CTkLabel(master = frame_left ,text = "Year :", justify = tkinter.LEFT).place(x=10,y=300)

a1 = customtkinter.CTkEntry(master = frame_left, width = 280 , placeholder_text = "First Name")
a1.place(x = 150,y = 50)
b1 = customtkinter.CTkEntry(master = frame_left, width = 280 , placeholder_text = "Last Name")
b1.place(x = 150,y = 100)
c1 = customtkinter.CTkEntry(master = frame_left, width = 280 , placeholder_text = "Email ID")
c1.place(x = 150,y = 150)
d1 = customtkinter.CTkEntry(master = frame_left, width = 280 , placeholder_text = "Contact Number without(+91)")
d1.place(x = 150,y = 200)
e1 = customtkinter.CTkOptionMenu(master=frame_left, values=["-Select-","CSE",'ECE','EEE','MECH',"CIVIL","AI & DS"], command=None)
e1.place(x=150,y=250)
f1 = customtkinter.CTkOptionMenu(master = frame_left ,values=["-Select-","I","II","III","IV"] )
f1.place(x=150,y=300)

excel()

btn = customtkinter.CTkButton(master = frame_left ,text="Submit", command=insert).place(x=150,y=350)

app.mainloop()
